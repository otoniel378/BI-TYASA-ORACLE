"""
scripts/download_snice_siderurgico.py — Descarga automática del reporte Excel
"Siderúrgico" (AUTOMÁTICOS > BENEFICIARIOS) desde el portal SNICE (snice.gob.mx).

El portal es una app JSF/Oracle ADF: las URLs de descarga viven dentro de una
sesión de navegador (jsessionid) y no son predecibles ni fijas, así que en vez
de requests.get() se navega con un navegador real (Playwright) y se captura el
evento de descarga al hacer click en el ícono de Excel de la fila "Siderúrgico",
en la columna del mes más reciente disponible (la más a la derecha).

Ruta de navegación (capturada con `playwright codegen` contra el sitio real —
todo el contenido vive dentro de un iframe, id "afr::DocWrapper"):
    Home -> botón "Acepto" (aviso de privacidad)
         -> link "Avisos y Permisos, Cupos y..." (tile de la portada)
         -> link "Avisos y Permisos" (página con 3 íconos: Cupos / Avisos y
            Permisos / Beneficiarios de Programas de Fomento)
         -> link "AUTOMÁTICOS"
         -> tabla BENEFICIARIOS (activa por default)

El archivo publicado va ~2 meses atrás del calendario real (ej. se descarga en
julio pero el "PERIODO REPORTADO" dentro del Excel dice mayo), así que el
archivo se nombra con el periodo REAL leído del propio Excel, no con la fecha
de descarga. Solo se conservan los 2 periodos más recientes en <out>/: al
llegar uno nuevo se borran los más viejos (el detalle no aporta más allá de
2 meses; el histórico agregado vive en Oracle, ver load_snice_to_oracle.py).

Uso:
    pip install playwright openpyxl
    playwright install chromium
    python scripts/download_snice_siderurgico.py
    python scripts/download_snice_siderurgico.py --headed           # ver el navegador (debug)
    python scripts/download_snice_siderurgico.py --out data/snice   # carpeta destino

Si el portal cambia su markup y esto deja de funcionar, volver a grabar con:
    playwright codegen --target python -o snice_codegen.py <SNICE_URL>
"""

import argparse
import re
import sys
from pathlib import Path

import openpyxl
from playwright.sync_api import sync_playwright

SNICE_URL = (
    "https://www.snice.gob.mx/AdminSNICE/faces/oracle/webcenter/portalapp/"
    "pages/paginasPublicas/publicHome.jspx"
)
IFRAME_SELECTOR = '[id="afr::DocWrapper"] iframe'
ROW_LABEL = "Siderúrgico"
NAV_TIMEOUT_MS = 45_000
DOWNLOAD_TIMEOUT_MS = 30_000
MAX_INTENTOS = 3
PERIODOS_A_CONSERVAR = 2

MESES_ES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10, "noviembre": 11,
    "diciembre": 12,
}


def _dismiss_privacy_notice(page):
    """
    Cierra el modal 'Aviso de privacidad simplificado' (botón 'Acepto').
    click() espera (auto-wait) a que el botón exista y sea clickeable en vez de
    solo revisar el DOM en el instante actual (con .count() había una carrera:
    en CI, más rápido/headless, el modal a veces aún no había renderizado
    cuando se revisaba, y el script seguía de largo dejándolo abierto y
    bloqueando todos los clics posteriores). No es fatal si nunca aparece.
    """
    try:
        page.get_by_role("button", name="Acepto").first.click(timeout=15_000)
        page.locator("#modalAvisoPrivacidad").wait_for(state="hidden", timeout=10_000)
    except Exception:
        pass


def _navigate_to_beneficiarios(page):
    """
    Home -> tile 'Avisos y Permisos, Cupos y...' -> link 'Avisos y Permisos'
    -> link 'AUTOMÁTICOS' -> tabla BENEFICIARIOS. Todo dentro del iframe
    afr::DocWrapper. Devuelve el FrameLocator ya posicionado en la tabla.
    """
    page.goto(SNICE_URL, wait_until="domcontentloaded")
    _dismiss_privacy_notice(page)

    frame = page.frame_locator(IFRAME_SELECTOR)

    frame.get_by_role(
        "link", name=re.compile(r"avisos y permisos,\s*cupos", re.I)
    ).first.click()

    frame.get_by_role("rowgroup").get_by_role(
        "link", name=re.compile(r"^avisos y permisos$", re.I)
    ).first.click()

    frame.get_by_role("link", name=re.compile("automáticos", re.I)).first.click()

    # Asegurar pestaña BENEFICIARIOS activa (suele venir activa por default)
    try:
        frame.locator("#panel-01 #myTabContent").click(timeout=5000)
    except Exception:
        pass

    frame.get_by_text(ROW_LABEL, exact=False).first.wait_for(
        state="visible", timeout=NAV_TIMEOUT_MS
    )
    return frame


def _last_excel_link(frame):
    """
    Ubica la fila '<ROW_LABEL>' y devuelve el link de descarga de su última
    columna (la más a la derecha = mes más reciente publicado).
    """
    row_label = frame.get_by_text(ROW_LABEL, exact=False).first
    row = row_label.locator("xpath=ancestor::tr[1]")
    if row.count() == 0:
        raise RuntimeError(f"No se pudo ubicar la fila de '{ROW_LABEL}'.")

    celdas = row.locator("td")
    if celdas.count() < 2:
        raise RuntimeError(f"La fila de '{ROW_LABEL}' no tiene columnas de meses.")

    ultima_celda = celdas.last
    icono = ultima_celda.locator("a").last
    if icono.count() == 0:
        raise RuntimeError(f"Sin ícono de descarga en la última columna de '{ROW_LABEL}'.")
    return icono


def _extraer_periodo(xlsx_path: Path) -> str:
    """
    Lee la celda 'PERIODO REPORTADO: <MES> DE <AÑO>' de la hoja AVISOS_AUTORIZADOS
    (primeras filas del encabezado) y la convierte a 'YYYY-MM'.
    """
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
    try:
        hoja = wb["AVISOS_AUTORIZADOS"] if "AVISOS_AUTORIZADOS" in wb.sheetnames else wb.worksheets[0]
        for row in hoja.iter_rows(min_row=1, max_row=10, max_col=1, values_only=True):
            texto = row[0]
            if texto and "PERIODO" in str(texto).upper():
                m = re.search(r"([A-ZÁÉÍÓÚÑ]+)\s+DE\s+(\d{4})", str(texto).upper())
                if m:
                    mes_num = MESES_ES.get(m.group(1).lower().replace("á", "a").replace("é", "e"))
                    if mes_num:
                        return f"{m.group(2)}-{mes_num:02d}"
        raise RuntimeError("No se encontró la celda 'PERIODO REPORTADO' en el Excel.")
    finally:
        wb.close()


def _aplicar_retencion(out_dir: Path, keep: int = PERIODOS_A_CONSERVAR):
    """Conserva solo los <keep> archivos siderurgico_YYYY-MM.xlsx más recientes."""
    archivos = sorted(out_dir.glob("siderurgico_*.xlsx"), reverse=True)
    for viejo in archivos[keep:]:
        viejo.unlink()
        print(f"  Retención: borrado {viejo.name}")


def download_siderurgico(out_dir: Path, headed: bool = False) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    last_error = None
    page = None

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=not headed)
        context = browser.new_context(accept_downloads=True)
        page = context.new_page()
        page.set_default_timeout(NAV_TIMEOUT_MS)

        for intento in range(1, MAX_INTENTOS + 1):
            try:
                frame = _navigate_to_beneficiarios(page)
                icono = _last_excel_link(frame)
                icono.scroll_into_view_if_needed()

                with context.expect_event("download", timeout=DOWNLOAD_TIMEOUT_MS) as download_info:
                    icono.click()
                download = download_info.value

                tmp_dest = out_dir / "_descarga_tmp.xlsx"
                download.save_as(str(tmp_dest))

                periodo = _extraer_periodo(tmp_dest)
                dest = out_dir / f"siderurgico_{periodo}.xlsx"
                tmp_dest.replace(dest)

                _aplicar_retencion(out_dir)

                browser.close()
                return dest

            except Exception as e:
                last_error = e
                print(f"  Intento {intento}/{MAX_INTENTOS} falló: {e}")

        # Todos los intentos fallaron: guardar screenshot para depurar antes de cerrar.
        try:
            debug_path = out_dir / "_debug_snice_failure.png"
            page.screenshot(path=str(debug_path), full_page=True)
            print(f"  Screenshot de depuración guardado en {debug_path}")
        except Exception:
            pass

        browser.close()
        raise RuntimeError(f"No se pudo descargar el reporte SNICE tras {MAX_INTENTOS} intentos: {last_error}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", default="data/snice", help="Carpeta destino")
    parser.add_argument("--headed", action="store_true", help="Correr con navegador visible (debug)")
    args = parser.parse_args()

    print(f"Descargando reporte SNICE '{ROW_LABEL}' (BENEFICIARIOS)...")
    try:
        dest = download_siderurgico(Path(args.out), headed=args.headed)
    except Exception as e:
        print(f"ERROR: {e}")
        sys.exit(1)

    print(f"OK: {dest}")


if __name__ == "__main__":
    main()
